import argparse
import openpyxl 
import os
import psutil
import time

def parse_args():
    parser = argparse.ArgumentParser(description='System resource usage monitor')
    parser.add_argument('--pid', '-p', type=int, help='PID')
    parser.add_argument('--cmd', type=str, help='command')
    return parser.parse_args()

def get_pid(cmd):
    with os.popen(cmd) as file_in:
        resp = file_in.readlines()
    if len(resp) < 2:
        return 0
    pid_of_myself = os.getpid()
    for r in resp[1:]:
        pid = int(resp[1].split()[0])
        if pid != pid_of_myself:
            return pid
    return 0

def main():
    args = parse_args()
    if args.pid is None:
        args.pid = get_pid(args.cmd)
        print(f'get_pid: {args.pid}')
    total_cpu_percent = 0
    total_memory_percent = 0
    count = 0
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1, value='CPU占用率')
    sheet.cell(row=1, column=2, value='平均CPU占用率')
    sheet.cell(row=1, column=3, value='内存占用率')
    sheet.cell(row=1, column=4, value='平均内存占用率')
    while True:
        try:
            p = psutil.Process(args.pid)
            cpu_percent = p.cpu_percent(interval=1)
            memory_percent = p.memory_percent()
            total_cpu_percent = total_cpu_percent + cpu_percent
            total_memory_percent = total_memory_percent + memory_percent
            count = count + 1
            average_cpu_percent = total_cpu_percent / count
            average_memory_percent = total_memory_percent / count
            sheet.cell(row=count+1, column=1, value=cpu_percent)
            sheet.cell(row=count+1, column=2, value=average_cpu_percent)
            sheet.cell(row=count+1, column=3, value=memory_percent)
            sheet.cell(row=count+1, column=4, value=average_memory_percent)
            print(f'CPU: ({cpu_percent}, {average_cpu_percent}), MEM: ({memory_percent}, {average_memory_percent})')
        except psutil.NoSuchProcess:
            break
    wb.save('monitor.xlsx')

if __name__ == '__main__':
    main()